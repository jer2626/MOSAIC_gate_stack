"""
MOSAIC Gate Stack Tier Recommendations — Excel Report
Generates a formatted Excel table, 2 recommendations per tier side by side.
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter

# ── Load data ─────────────────────────────────────────────────────────────────
df = pd.read_csv('D_gate_stack_database.csv')
feas = df[df['feasible'] == True].copy()

def _s(v):
    sub = str.maketrans('₀₁₂₃₄₅₆₇₈₉', '0123456789')
    return str(v).translate(sub).encode('ascii', 'replace').decode('ascii')

def apply_filters(d, req):
    if 'min_ION' in req: d = d[d['ION_uAum'] >= req['min_ION']]
    if 'max_IG'  in req: d = d[d['IG_Aum']   <= req['max_IG']]
    if 'max_SS'  in req: d = d[d['SS_mVdec'] <= req['max_SS']]
    if 'max_EOT' in req: d = d[d['EOT_nm']   <= req['max_EOT']]
    if 'min_VTH' in req: d = d[d['VTH_V']    >= req['min_VTH']]
    if 'max_VTH' in req: d = d[d['VTH_V']    <= req['max_VTH']]
    if 'max_dT'  in req: d = d[d['dT_K']     <= req['max_dT']]
    if 'min_VBD' in req: d = d[d['VBD_V']    >= req['min_VBD']]
    if 'max_Eox_frac_EBD' in req:
        d = d[d['Eox_MVcm'] <= req['max_Eox_frac_EBD'] * d['EBD_HK']]
    return d.copy()

def composite_score(d):
    return (
        np.log10(d['ION_uAum'].clip(1))      * 0.35 +
        (-np.log10(d['IG_Aum'].clip(1e-30))) * 0.25 +
        (-d['SS_mVdec'] / 60)                * 0.20 +
        d['Rel_score']                        * 0.20
    )

TIERS = {
    'Tier 1: Foundational Design':      {'min_ION':100,  'max_IG':1e-6, 'max_SS':90, 'max_EOT':1.5, 'min_VTH':0.3, 'max_VTH':0.7},
    'Tier 2: Balanced Performance':     {'min_ION':300,  'max_IG':1e-7, 'max_SS':80, 'max_EOT':1.0, 'min_VTH':0.3, 'max_VTH':0.6},
    'Tier 3: Reliability + Thermal':    {'min_ION':250,  'max_IG':1e-7, 'max_SS':75, 'max_EOT':0.9, 'max_Eox_frac_EBD':0.70, 'min_VBD':1.2, 'max_dT':20},
    'Tier 4: Pareto + Electro-Thermal': {'min_ION':300,  'max_IG':1e-8, 'max_SS':70, 'max_EOT':0.7, 'max_Eox_frac_EBD':0.65, 'max_dT':15,  'min_VBD':1.2},
}

TIER_COLORS = {
    'Tier 1: Foundational Design':      '1A6FAF',
    'Tier 2: Balanced Performance':     '2E8B57',
    'Tier 3: Reliability + Thermal':    'C25F00',
    'Tier 4: Pareto + Electro-Thermal': '8B1A1A',
}

CONSTRAINTS_TEXT = {
    'Tier 1: Foundational Design':      'ION≥1e2 μA/μm | IG≤1e-6 A/μm | SS≤90 mV/dec | EOT≤1.5 nm | VTH 0.3–0.7 V',
    'Tier 2: Balanced Performance':     'ION≥3e2 μA/μm | IG≤1e-7 A/μm | SS≤80 mV/dec | EOT≤1.0 nm | VTH 0.3–0.6 V',
    'Tier 3: Reliability + Thermal':    'ION≥2.5e2 μA/μm | IG≤1e-7 A/μm | SS≤75 mV/dec | EOT≤0.9 nm | Eox≤70% EBD | VBD≥1.2 V | ΔT≤20 K',
    'Tier 4: Pareto + Electro-Thermal': 'ION≥3e2 μA/μm | IG≤1e-8 A/μm | SS≤70 mV/dec | EOT≤0.7 nm | Eox≤65% EBD | ΔT≤15 K | VBD≥1.2 V',
}

# Row definitions: (label, unit, field, format_func)
ROWS = [
    # Section: Gate Stack
    ('GATE STACK', None, None, None),
    ('Gate Electrode', '',      'Gate',        lambda r: _s(r['Gate'])),
    ('Dielectric',     '',      'Dielectric',  lambda r: _s(r['Dielectric'])),
    ('Substrate',      '',      'Substrate',   lambda r: str(r['Substrate'])),
    ('Interfacial Layer', '',   't_IL_nm',     lambda r: f"SiOx, {r['t_IL_nm']:.2f} nm" if r['t_IL_nm'] > 0 else 'None'),
    # Section: Structural Parameters
    ('STRUCTURAL PARAMETERS', None, None, None),
    ('HK Thickness (t_HK)', 'nm',    't_HK_nm', lambda r: f"{r['t_HK_nm']:.2f}"),
    ('Gate Length (L)',      'nm',    'L_nm',    lambda r: f"{r['L_nm']:.0f}"),
    ('Supply Voltage (VDD)', 'V',     'VDD',     lambda r: f"{r['VDD']:.2f}"),
    ('Channel Doping (Nsub)','cm⁻³', 'Nsub', lambda r: f"{r['Nsub']:.2e}"),
    # Section: Performance
    ('PERFORMANCE', None, None, None),
    ('Drive Current (ION)',   'μA/μm', 'ION_uAum',  lambda r: f"{r['ION_uAum']:.1f}"),
    ('Subthreshold Slope (SS)', 'mV/dec', 'SS_mVdec', lambda r: f"{r['SS_mVdec']:.1f}"),
    ('Threshold Voltage (VTH)', 'V',    'VTH_V',   lambda r: f"{r['VTH_V']:.3f}"),
    ('Equiv. Oxide Thickness (EOT)', 'nm', 'EOT_nm', lambda r: f"{r['EOT_nm']:.3f}"),
    # Section: Reliability
    ('RELIABILITY', None, None, None),
    ('Gate Leakage (IG)',    'A/μm',  'IG_Aum',    lambda r: f"{r['IG_Aum']:.2e}"),
    ('Oxide Field (Eox)',    'MV/cm',  'Eox_MVcm',  lambda r: f"{r['Eox_MVcm']:.2f}"),
    ('Breakdown Voltage (VBD)', 'V',  'VBD_V',     lambda r: f"{r['VBD_V']:.2f}"),
    ('Self-Heating (ΔT)', 'K',   'dT_K',      lambda r: f"{r['dT_K']:.1f}"),
    ('Reliability Score',   '0–1', 'Rel_score', lambda r: f"{r['Rel_score']:.3f}"),
    ('Composite Score',     '',       '_score',    lambda r: f"{r['_score']:.3f}"),
]

SECTION_ROWS = {'GATE STACK', 'STRUCTURAL PARAMETERS', 'PERFORMANCE', 'RELIABILITY'}

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Tier Recommendations'

# Helper styles
def make_fill(hex_color, tint=False):
    if tint:
        # lighter version: blend with white
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        r2 = int(r + (255-r)*0.82)
        g2 = int(g + (255-g)*0.82)
        b2 = int(b + (255-b)*0.82)
        hex2 = f'{r2:02X}{g2:02X}{b2:02X}'
        return PatternFill('solid', fgColor=hex2)
    return PatternFill('solid', fgColor=hex_color)

thin  = Side(style='thin',   color='CCCCCC')
thick = Side(style='medium', color='999999')

def border(left=thin, right=thin, top=thin, bottom=thin):
    return Border(left=left, right=right, top=top, bottom=bottom)

def cell(ws, row, col, value='', bold=False, color='000000', bg=None,
         align='left', size=10, wrap=False, italic=False, border_obj=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=size, italic=italic,
                  name='Calibri')
    c.alignment = Alignment(horizontal=align, vertical='center',
                             wrap_text=wrap)
    if bg:
        c.fill = make_fill(bg)
    if border_obj:
        c.border = border_obj
    return c

# ── Layout constants ──────────────────────────────────────────────────────────
# Columns: A=parameter label, B=unit, then pairs per tier (Rank1 | Rank2)
# Col layout: 1=Label, 2=Unit, 3=T1R1, 4=T1R2, 5=T2R1, 6=T2R2,
#             7=T3R1, 8=T3R2, 9=T4R1, 10=T4R2
LABEL_COL = 1
UNIT_COL  = 2
TIER_START_COLS = [3, 5, 7, 9]   # start col for each tier (2 cols each)

# Column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 10
for ci in range(3, 11):
    ws.column_dimensions[get_column_letter(ci)].width = 18

# ── Title row ─────────────────────────────────────────────────────────────────
ROW = 1
ws.row_dimensions[ROW].height = 28
cell(ws, ROW, 1, 'MOSAIC Gate Stack Recommendations — Top 2 Designs per Tier (Database Search)',
     bold=True, color='FFFFFF', bg='1A1A2E', align='center', size=12,
     border_obj=border())
ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=10)

# ── Tier headers ──────────────────────────────────────────────────────────────
ROW = 2
ws.row_dimensions[ROW].height = 22
tier_items = list(TIERS.items())
for ti, (tier_name, req) in enumerate(tier_items):
    sc = TIER_START_COLS[ti]
    color = TIER_COLORS[tier_name]
    n_matched = len(apply_filters(feas, req))
    cell(ws, ROW, sc, tier_name.split(':')[0],
         bold=True, color='FFFFFF', bg=color, align='center', size=11,
         border_obj=border())
    cell(ws, ROW, sc+1, f'({n_matched} matched)',
         bold=False, color='FFFFFF', bg=color, align='center', size=9,
         italic=True, border_obj=border())

# Label/unit header
cell(ws, ROW, LABEL_COL, 'Parameter',
     bold=True, color='FFFFFF', bg='2C2C2C', align='left', size=10,
     border_obj=border())
cell(ws, ROW, UNIT_COL, 'Unit',
     bold=True, color='FFFFFF', bg='2C2C2C', align='center', size=10,
     border_obj=border())

# ── Constraints row ───────────────────────────────────────────────────────────
ROW = 3
ws.row_dimensions[ROW].height = 16
for ti, (tier_name, _) in enumerate(tier_items):
    sc = TIER_START_COLS[ti]
    color = TIER_COLORS[tier_name]
    txt = CONSTRAINTS_TEXT[tier_name]
    cell(ws, ROW, sc, txt,
         bold=False, color=color, bg='F8F8F8', align='center', size=7,
         wrap=True, border_obj=border())
    ws.merge_cells(start_row=ROW, start_column=sc, end_row=ROW, end_column=sc+1)

cell(ws, ROW, LABEL_COL, 'Constraints',
     bold=True, color='555555', bg='F0F0F0', align='left', size=9,
     italic=True, border_obj=border())
cell(ws, ROW, UNIT_COL, '',
     bg='F0F0F0', border_obj=border())

# ── Rank sub-headers ──────────────────────────────────────────────────────────
ROW = 4
ws.row_dimensions[ROW].height = 16
cell(ws, ROW, LABEL_COL, '', bg='EFEFEF', border_obj=border())
cell(ws, ROW, UNIT_COL,  '', bg='EFEFEF', border_obj=border())
for ti, (tier_name, _) in enumerate(tier_items):
    sc = TIER_START_COLS[ti]
    color = TIER_COLORS[tier_name]
    for ri, rank_lbl in enumerate(['Rank 1', 'Rank 2']):
        cell(ws, ROW, sc+ri, rank_lbl,
             bold=True, color=color, bg='EFEFEF', align='center', size=9,
             border_obj=border())

# ── Get top 2 per tier ────────────────────────────────────────────────────────
top2_per_tier = []
for tier_name, req in tier_items:
    d = apply_filters(feas, req)
    d['_score'] = composite_score(d)
    top2 = d.sort_values('_score', ascending=False).head(2).reset_index(drop=True)
    top2_per_tier.append(top2)

# ── Data rows ─────────────────────────────────────────────────────────────────
ROW = 5
alt = False
for label, unit, field, fmt_fn in ROWS:
    ws.row_dimensions[ROW].height = 15 if label not in SECTION_ROWS else 14

    if label in SECTION_ROWS:
        # Section header spanning all columns
        cell(ws, ROW, LABEL_COL, label,
             bold=True, color='FFFFFF', bg='444444', align='left', size=9,
             border_obj=border())
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=10)
        ROW += 1
        alt = False
        continue

    bg_row = 'F5F8FF' if alt else 'FFFFFF'
    alt = not alt

    # Label cell
    cell(ws, ROW, LABEL_COL, label,
         bold=False, color='222222', bg=bg_row, align='left', size=9,
         border_obj=border())
    # Unit cell
    cell(ws, ROW, UNIT_COL, unit or '',
         bold=False, color='666666', bg=bg_row, align='center', size=9,
         italic=True, border_obj=border())

    # Values per tier
    for ti, top2 in enumerate(top2_per_tier):
        sc = TIER_START_COLS[ti]
        color = TIER_COLORS[tier_items[ti][0]]
        for ri in range(2):
            if ri < len(top2):
                r = top2.iloc[ri]
                val = fmt_fn(r)
                is_depl = (field == 'VTH_V' and r['VTH_V'] < 0)
                txt_color = 'CC0000' if is_depl else '1A1A1A'
                bold_val  = is_depl
                cell(ws, ROW, sc+ri, val,
                     bold=bold_val, color=txt_color, bg=bg_row,
                     align='center', size=9,
                     border_obj=border())
            else:
                cell(ws, ROW, sc+ri, 'N/A',
                     color='999999', bg=bg_row, align='center', size=9,
                     border_obj=border())
    ROW += 1

# ── Notes row ─────────────────────────────────────────────────────────────────
ws.row_dimensions[ROW].height = 14
note = ('Scoring: ION 35% · IG 25% · SS 20% · Rel 20%  |  '
        'Red VTH = depletion-mode (VTH < 0, device on at zero gate bias)  |  '
        'Source: D_gate_stack_database.csv  |  NMOS, W = 1 μm')
cell(ws, ROW, 1, note,
     bold=False, color='888888', bg='F8F8F8', align='left', size=7.5,
     italic=True, border_obj=border())
ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=10)

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = 'C5'

# ── Save ─────────────────────────────────────────────────────────────────────
out = 'MOSAIC_Tier_Recommendations.xlsx'
wb.save(out)
print(f'Saved: {out}')
